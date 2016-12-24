import argparse
import os
import sys
import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook

desired_width = 179
pd.set_option('display.width', desired_width)

'''
Header Fields
Listing Number,Status,Property Sub-Type,Street #,Street Name,
Original List Price,Listing Price,Sold Price,Cumulative DOM,
Days On Market,# Bedrooms,Baths - Total,Baths - Full,Baths - 3/4,
Baths - 1/2,Baths - 1/4,Approx SqFt,Lot SqFt,Year Built,Sold Date
'''


def get_active(df, status='A', prtype='SF'):
    """extracts the Active Single Family Listings from the DataFrame
    :param status:  A=Active, P= Pending C=Sold
    :param Prtype:  SF=Single Family, CN= Condo, TH= Town home
    """
    if prtype == 'SF':
        selected = df[(df.Status == status) & (df.Prtype == prtype)]
    else:
        sel = df[((df.Status == status) & (df.Prtype == 'CN')) |
                 ((df.Status == status) & (df.Prtype == 'TH'))]
        selected = sel.sort_values(by=['Status', 'Sale Price', 'Listing Price'])

    return selected


def get_sold(df, prtype='SF'):
    '''
    Extracts the Sold listings from the DataFrame
    :param df: input DataFrame
    :param prtype: Property type SF= Single Family, CN= Condos, TH= Town Home
    :return: Extracted DataFrame
    '''
    if prtype == 'SF':
        selected = df[(df.Status == 'C') & (df.Prtype == 'SF')]
    else:
        sel = df[((df.Status == 'C') & (df.Prtype == 'CN')) |
                 ((df.Status == 'C') & (df.Prtype == 'TH'))]
        selected = sel.sort_values(by=['Status', 'Sale Price', 'Listing Price'])

    return selected


def write_sheet(list_df, path_name):
    '''
    Writes the given sheet name to the open workbook
    :param list_df:  List of DataFrame to write to sheets of the open workbook
    :param path_name: path to the workbook
    :return: None
    '''
    '''def save_xls(list_dfs, xls_path):
    writer = ExcelWriter(xls_path)
    for n, df in enumerate(list_dfs):
        df.to_excel(writer,'sheet%s' % n)
    writer.save()'''

    tday = datetime.datetime.today()
    '''if os.path.exists('test.xlsx') == False:
        book = load_workbook('test.xlsx')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        current = tday.strftime('%y-%m-%d-%H-%M-%S')
        print('current= ', current)
        book.create_sheet(current, 0)
    else:'''
    try:
        with pd.ExcelWriter(path_name, engine='openpyxl') as writer:

            current = tday.strftime('%y-%m-%d')
            for name, df in list_df:
                df.to_excel(writer,
                            sheet_name=str.format('{}{}', name, current))
            writer.save()
    except IOError as e:
        errno, strerror = e.args
        print(
            "I/O error({0}): {1} You probably need to close the open Excel file {2}".format(
                errno, strerror, path_name))
    except:
        print("Unexpected error:", sys.exc_info()[0])
        raise


def do_convert(file_name):
    file = 'shores.xlsx' if file_name is None else file_name
    root, ext = os.path.splitext(file)
    if ext == '.csv':
        df = pd.read_csv(file)
    elif ext == '.xlsx':
        df = pd.read_excel(file)
    else:
        return None

    df['Street #'] = df['Street #'].map(str) + " " + df['Street Name']

    for index, row in df.iterrows():
        row['Baths - Total'] = row['Baths - Full'] + row['Baths - 3/4'] * .75 + \
                               row['Baths - 1/2'] * .5 + row[
                                                             'Baths - 1/4'] * .25

    if args.verbose: print(df.head())

    df.drop(['Listing Number', 'Street Name', 'Days On Market', 'Baths - Full',
             'Baths - 3/4', 'Baths - 1/2', 'Baths - 1/4'], axis=1, inplace=True)

    df.rename(
        columns={'Property Sub-Type': 'Prtype', 'Street #': 'Sold Listing',
                 '# Bedrooms': 'Bed',
                 'Baths - Total': 'Bath', 'SqFt': 'Interior Size',
                 'Year Built': 'Yr. Built', 'Lot SqFt': 'Lot Size',
                 'Sold Price': 'Sale Price', 'Cumulative DOM': 'CDOM'},
        inplace=True)
    cols = df.columns.tolist()

    if args.verbose: print(df.head())

    if isinstance(df['Sold Date'][0], str):
        df['Sold Date'] = df['Sold Date'].apply(lambda x:
                '' if pd.isnull(x) else datetime.datetime.strptime(
                x, '%Y-%m-%d').strftime('%m/%d/%y'))
    else:
        df['Sold Date'] = df['Sold Date'].apply(
            lambda x: '' if pd.isnull(x) else x.strftime('%m/%d/%y'))

    if args.verbose: print(df.head())

    # 	Status = [ A: Active, C: Sold, P: Pending]
    df.sort_values(by=['Status', 'Prtype', 'Sale Price', 'Listing Price'],
                   inplace=True)

    sheets = []
    sheets.append(('Active_SFR', get_active(df)))
    sheets.append(('Escrow_SFR', get_active(df, 'P')))
    sheets.append(('Sold_SFR', get_sold(df)))
    sheets.append(('Active_Condos', get_active(df, 'A', 'CN')))
    sheets.append(('Escrow_Condos', get_active(df, 'P', 'CN')))
    sheets.append(('Sold_Condos', get_sold(df, 'CN')))

    write_sheet(sheets, 'test.xlsx')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("infile",
                        help="the input file to use to generate the Activity text")
    parser.add_argument("outfile",
                        help="the output file containing the Activity text")
    parser.add_argument("--verbose", help="increase output verbosity",
                        action="store_true")
    args = parser.parse_args()
    if args.verbose:
        print("verbosity turned on")

    do_convert(args.infile)
